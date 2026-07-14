"""酒店经营日报导入公共接口的行为测试。"""

import sqlite3  # 构造旧版数据库，验证真实迁移而不是新建路径。
from datetime import date  # 使用明确营业日调用公共查询接口。
from pathlib import Path  # 为测试生成隔离的日报文件和数据目录。

import pytest  # 验证无效日报会通过公共接口返回明确错误。
from openpyxl import Workbook  # 生成最小 XLSX，覆盖真实 Excel 日期行为。

from enterprise.hotel import HotelDataWorkspace  # 只测试用户确认的公共边界。


def _create_legacy_report_database(root: Path, mapping_json: str | None) -> None:
    """创建没有映射快照列和迁移台账的旧版日报数据库。

    Args:
        root: 旧版数据库所在的本地数据目录。
        mapping_json: 旧映射表内容；``None`` 表示模拟损坏的映射表结构。

    Returns:
        None.
    """

    # 创建旧应用已经使用的数据目录。
    root.mkdir()
    # 根据测试场景决定旧映射表是否拥有可回填的 JSON 字段。
    mapping_column = "mapping_json TEXT NOT NULL," if mapping_json is not None else ""
    # 使用真实 SQLite 文件作为公开构造函数的升级输入。
    with sqlite3.connect(root / "enterprise.db") as connection:
        # 旧日报版本表故意不包含 mapping_json 和迁移台账。
        connection.executescript(
            f"""
            CREATE TABLE hotel_report_mappings (
                template_name TEXT PRIMARY KEY,
                {mapping_column}
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            CREATE TABLE hotel_report_versions (
                id TEXT PRIMARY KEY,
                template_name TEXT NOT NULL,
                business_date TEXT NOT NULL,
                version INTEGER NOT NULL,
                records_json TEXT NOT NULL,
                source_name TEXT NOT NULL,
                stored_path TEXT NOT NULL,
                sha256 TEXT NOT NULL,
                imported_at TEXT NOT NULL,
                is_active INTEGER NOT NULL,
                UNIQUE(template_name, business_date, version)
            );
            """
        )
        # 正常旧库保存模板 JSON；损坏场景只保存名称和时间。
        mapping_values = (
            ("pms-daily", mapping_json, "2026-07-13T08:00:00+00:00", "2026-07-13T08:00:00+00:00")
            if mapping_json is not None
            else ("pms-daily", "2026-07-13T08:00:00+00:00", "2026-07-13T08:00:00+00:00")
        )
        # 占位符数量与上面的旧表结构保持一致。
        placeholders = "?, ?, ?, ?" if mapping_json is not None else "?, ?, ?"
        connection.execute(
            f"INSERT INTO hotel_report_mappings VALUES ({placeholders})",
            mapping_values,
        )
        # 旧历史日报用于验证升级、失败重试和映射回填。
        connection.execute(
            "INSERT INTO hotel_report_versions VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                "report_legacy",
                "pms-daily",
                "2026-07-13",
                1,
                '[{"business_date":"2026-07-13","rooms_sold":"150"}]',
                "legacy.csv",
                "hotel_imports/report_legacy/legacy.csv",
                "legacy-sha256",
                "2026-07-13T08:00:00+00:00",
                1,
            ),
        )


def test_saved_mapping_imports_csv_by_business_date(tmp_path: Path) -> None:
    """负责人保存字段映射后，可导入日报并按营业日读取标准化数据。

    Args:
        tmp_path: Pytest 为本测试提供的隔离临时目录。

    Returns:
        None.
    """

    # 创建独立工作空间，确保测试不读取真实酒店数据。
    workspace = HotelDataWorkspace(tmp_path / "enterprise")
    # 保存负责人首次确认的 PMS 来源表头映射。
    workspace.save_mapping(
        "pms-daily",
        {
            "business_date": "营业日期",
            "available_rooms": "可售房",
            "rooms_sold": "已售房",
            "room_revenue": "客房收入",
        },
    )
    # 写入一份只含单一营业日的最小真实 CSV 日报。
    report_path = tmp_path / "pms-daily.csv"
    report_path.write_text(
        "营业日期,可售房,已售房,客房收入\n2026-07-13,200,150,90000\n",
        encoding="utf-8",
    )

    # 通过公共导入接口生成第一个日报版本。
    imported = workspace.import_report(report_path, "pms-daily")
    # 模拟驾驶舱按营业日读取当前生效数据。
    active = workspace.get_active_report(date(2026, 7, 13), "pms-daily")

    # 营业日必须来自报表内容，而不是文件名或导入时间。
    assert imported.business_date == date(2026, 7, 13)
    # 首次导入版本固定从一开始。
    assert imported.version == 1
    # 来源表头应被替换为稳定的标准字段名。
    assert imported.records == [
        {
            "business_date": "2026-07-13",
            "available_rooms": "200",
            "rooms_sold": "150",
            "room_revenue": "90000",
        }
    ]
    # 按营业日读取到的对象应与刚导入的生效版本完全一致。
    assert active == imported


def test_reimport_keeps_history_and_activates_latest_version(tmp_path: Path) -> None:
    """同一营业日重复导入会保留旧数据，并让最新版本参与后续读取。

    Args:
        tmp_path: Pytest 为本测试提供的隔离临时目录。

    Returns:
        None.
    """

    # 为修订场景创建独立工作空间和固定字段映射。
    workspace = HotelDataWorkspace(tmp_path / "enterprise")
    workspace.save_mapping(
        "pms-daily",
        {
            "business_date": "营业日期",
            "available_rooms": "可售房",
            "rooms_sold": "已售房",
            "room_revenue": "客房收入",
        },
    )
    # 第一份文件代表夜审后最初导出的日报。
    first_path = tmp_path / "pms-first.csv"
    first_path.write_text(
        "营业日期,可售房,已售房,客房收入\n2026-07-13,200,150,90000\n",
        encoding="utf-8",
    )
    # 第二份文件代表负责人发现差异后的修订日报。
    second_path = tmp_path / "pms-corrected.csv"
    second_path.write_text(
        "营业日期,可售房,已售房,客房收入\n2026-07-13,200,160,96000\n",
        encoding="utf-8",
    )

    # 先导入原始版本，再导入同营业日修订版本。
    workspace.import_report(first_path, "pms-daily")
    corrected = workspace.import_report(second_path, "pms-daily")
    # 只通过公共列表接口读取完整版本历史。
    versions = workspace.list_report_versions(date(2026, 7, 13), "pms-daily")

    # 历史必须按最新到最旧排列且版本连续递增。
    assert [item.version for item in versions] == [2, 1]
    # 两份已售房数据都要保留，不能静默覆盖旧版本。
    assert [item.records[0]["rooms_sold"] for item in versions] == ["160", "150"]
    # 驾驶舱当前读取必须自动指向修订后的版本。
    assert workspace.get_active_report(date(2026, 7, 13), "pms-daily") == corrected


def test_import_rejects_report_missing_a_mapped_source_column(tmp_path: Path) -> None:
    """报表缺少模板声明的字段时必须阻断，不能静默保存空值。

    Args:
        tmp_path: Pytest 为本测试提供的隔离临时目录。

    Returns:
        None.
    """

    # 模板声明已售房是后续计算必需的来源字段。
    workspace = HotelDataWorkspace(tmp_path / "enterprise")
    workspace.save_mapping(
        "pms-daily",
        {
            "business_date": "营业日期",
            "available_rooms": "可售房",
            "rooms_sold": "已售房",
        },
    )
    # 实际文件故意遗漏已售房列，模拟 PMS 导出模板变化。
    report_path = tmp_path / "missing-column.csv"
    report_path.write_text(
        "营业日期,可售房\n2026-07-13,200\n",
        encoding="utf-8",
    )

    # 错误必须指出具体缺失表头，帮助负责人修复映射。
    with pytest.raises(ValueError, match="缺少已映射字段: 已售房"):
        workspace.import_report(report_path, "pms-daily")


def test_saved_mapping_is_reused_for_excel_after_restart(tmp_path: Path) -> None:
    """字段映射持久化后，重新打开工作空间仍可导入 Excel 日报。

    Args:
        tmp_path: Pytest 为本测试提供的隔离临时目录。

    Returns:
        None.
    """

    # 首次启动时保存映射，随后丢弃该工作空间对象。
    root = tmp_path / "enterprise"
    workspace = HotelDataWorkspace(root)
    workspace.save_mapping(
        "pms-daily",
        {
            "business_date": "营业日期",
            "available_rooms": "可售房",
            "rooms_sold": "已售房",
        },
    )
    # 使用 openpyxl 写入真正的 Excel 日期和数字单元格。
    report_path = tmp_path / "pms-daily.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["营业日期", "可售房", "已售房"])
    worksheet.append([date(2026, 7, 13), 200, 150])
    workbook.save(report_path)
    workbook.close()

    # 重新创建工作空间，证明映射来自持久化数据而非内存。
    reopened = HotelDataWorkspace(root)
    imported = reopened.import_report(report_path, "pms-daily")

    # Excel 日期应规范为 ISO 文本，数字单元格仍保持数字类型。
    assert imported.records == [
        {"business_date": "2026-07-13", "available_rooms": 200, "rooms_sold": 150}
    ]
    # 持久化后重新读取的版本必须与导入返回值一致。
    assert reopened.get_active_report(date(2026, 7, 13), "pms-daily") == imported


def test_import_rejects_conflicting_business_dates(tmp_path: Path) -> None:
    """一个日报包含多个营业日时，应列出冲突日期并停止导入。

    Args:
        tmp_path: Pytest 为本测试提供的隔离临时目录。

    Returns:
        None.
    """

    # 创建只需要营业日和已售房的最小模板。
    workspace = HotelDataWorkspace(tmp_path / "enterprise")
    workspace.save_mapping(
        "pms-daily",
        {"business_date": "营业日期", "rooms_sold": "已售房"},
    )
    # 文件故意混入两个营业日，模拟错误拼接的 PMS 日报。
    report_path = tmp_path / "conflicting-dates.csv"
    report_path.write_text(
        "营业日期,已售房\n2026-07-13,150\n2026-07-14,160\n",
        encoding="utf-8",
    )

    # 阻断信息必须稳定列出两个冲突日期供负责人确认。
    with pytest.raises(
        ValueError,
        match="发现多个营业日: 2026-07-13、2026-07-14",
    ):
        workspace.import_report(report_path, "pms-daily")


def test_report_version_preserves_the_mapping_used_at_import(tmp_path: Path) -> None:
    """模板后续变更时，历史日报仍保留导入当时的字段映射。

    Args:
        tmp_path: Pytest 为本测试提供的隔离临时目录。

    Returns:
        None.
    """

    # 保存第一版模板并导入与之匹配的历史日报。
    workspace = HotelDataWorkspace(tmp_path / "enterprise")
    original_mapping = {"business_date": "营业日期", "rooms_sold": "已售房"}
    workspace.save_mapping("pms-daily", original_mapping)
    report_path = tmp_path / "pms-daily.csv"
    report_path.write_text(
        "营业日期,已售房\n2026-07-13,150\n",
        encoding="utf-8",
    )
    workspace.import_report(report_path, "pms-daily")

    # 模拟 PMS 升级后来源表头发生变化并更新当前模板。
    workspace.save_mapping(
        "pms-daily",
        {"business_date": "业务日期", "rooms_sold": "出租房"},
    )
    # 通过公共版本接口重新读取旧日报。
    historical = workspace.list_report_versions(date(2026, 7, 13), "pms-daily")[0]

    # 旧日报必须冻结原映射，不能被当前模板重新解释。
    assert historical.mapping == original_mapping


def test_import_accepts_only_csv_or_xlsx_reports(tmp_path: Path) -> None:
    """经营日报入口只接受已确认的 CSV/XLSX 文件类型。

    Args:
        tmp_path: Pytest 为本测试提供的隔离临时目录。

    Returns:
        None.
    """

    # 即使文本内容看起来像表格，扩展名仍不属于产品承诺范围。
    workspace = HotelDataWorkspace(tmp_path / "enterprise")
    workspace.save_mapping("pms-daily", {"business_date": "营业日期"})
    report_path = tmp_path / "pms-daily.txt"
    report_path.write_text("营业日期\n2026-07-13\n", encoding="utf-8")

    # 错误信息应明确说明首版接受的两种格式。
    with pytest.raises(ValueError, match="经营日报只支持 CSV 或 XLSX 文件"):
        workspace.import_report(report_path, "pms-daily")


def test_excel_import_ignores_sheets_that_do_not_match_mapping(tmp_path: Path) -> None:
    """多工作表 Excel 可忽略封面或说明页，只导入匹配模板的数据表。

    Args:
        tmp_path: Pytest 为本测试提供的隔离临时目录。

    Returns:
        None.
    """

    # 保存只匹配实际经营数据表的字段映射。
    workspace = HotelDataWorkspace(tmp_path / "enterprise")
    workspace.save_mapping(
        "pms-daily",
        {"business_date": "营业日期", "rooms_sold": "已售房"},
    )
    # 构造包含说明页和经营数据页的常见 PMS 工作簿。
    report_path = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    cover = workbook.active
    cover.title = "说明"
    cover.append(["PMS 日报导出说明"])
    cover.append(["以下工作表才是经营数据"])
    data_sheet = workbook.create_sheet("经营日报")
    data_sheet.append(["营业日期", "已售房"])
    data_sheet.append([date(2026, 7, 13), 150])
    workbook.save(report_path)
    workbook.close()

    # 导入公共接口应自动选择完整匹配映射的数据工作表。
    imported = workspace.import_report(report_path, "pms-daily")

    # 说明页内容不得混入标准化经营记录。
    assert imported.records == [{"business_date": "2026-07-13", "rooms_sold": 150}]


def test_legacy_report_schema_is_migrated_without_losing_history(tmp_path: Path) -> None:
    """旧版日报表升级后仍可读取，并补齐导入时使用的映射快照。

    Args:
        tmp_path: Pytest 为本测试提供的隔离临时目录。

    Returns:
        None.
    """

    # 创建包含模板和历史日报的正常旧版数据库。
    root = tmp_path / "enterprise"
    _create_legacy_report_database(
        root,
        '{"business_date":"营业日期","rooms_sold":"已售房"}',
    )

    # 正常打开公共工作空间应自动、事务化完成缺失迁移。
    workspace = HotelDataWorkspace(root)
    # 迁移结果仍通过既定版本列表接口验证，不直接查询新表结构。
    historical = workspace.list_report_versions(date(2026, 7, 13), "pms-daily")[0]

    # 旧版标准记录必须原样保留。
    assert historical.records[0]["rooms_sold"] == "150"
    # 新增映射快照列应使用旧版模板完成回填。
    assert historical.mapping == {"business_date": "营业日期", "rooms_sold": "已售房"}


def test_failed_schema_migration_rolls_back_and_can_be_retried(tmp_path: Path) -> None:
    """迁移中途失败不会半提交，修复旧库后可完整重试。

    Args:
        tmp_path: Pytest 为本测试提供的隔离临时目录。

    Returns:
        None.
    """

    # 构造缺失 mapping_json 的损坏旧映射表，使版本二回填必然失败。
    root = tmp_path / "enterprise"
    _create_legacy_report_database(root, None)

    # 第一次启动必须暴露真实数据库错误，而不是记录一个虚假的成功版本。
    with pytest.raises(sqlite3.OperationalError, match="mapping_json"):
        HotelDataWorkspace(root)

    # 模拟负责人使用修复工具补回旧映射表缺少的 JSON 列和值。
    with sqlite3.connect(root / "enterprise.db") as connection:
        connection.execute(
            "ALTER TABLE hotel_report_mappings ADD COLUMN mapping_json TEXT NOT NULL DEFAULT '{}'"
        )
        connection.execute(
            "UPDATE hotel_report_mappings SET mapping_json=? WHERE template_name=?",
            ('{"business_date":"营业日期","rooms_sold":"已售房"}', "pms-daily"),
        )

    # 再次正常启动必须重跑完整迁移，而不是沿用第一次的半成品。
    workspace = HotelDataWorkspace(root)
    historical = workspace.list_report_versions(date(2026, 7, 13), "pms-daily")[0]

    # 正确回填说明 ALTER TABLE 和迁移台账在失败时都已回滚。
    assert historical.mapping == {"business_date": "营业日期", "rooms_sold": "已售房"}
